# HM.py — HM – Processador Hospitalar TXT → Excel
# VERSÃO DEFINITIVA 100% CLOUD (COM FÓRMULAS REAIS NO EXCEL)
# AJUSTES: FORMATAÇÃO NUMÉRICA COMPLETA

import io
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

st.set_page_config(page_title="HM — Processador Hospitalar",
                   page_icon="🏥", layout="wide")

FIXED_HEADERS_A_TO_Y = [
    "REGISTRO", "NOME DO PACIENTE", "ENTRADA", "SAÍDA", "TIPO DE PRODUTO",
    "CÓD. PRODUTO", "DESC. PRODUTO", "QUANTIDADE", "COMANDA", "CÓD. TUSS",
    "DESTINO", "DATA E HORA DO PROC.", "UNIDADE", "MÉDICO", "SETOR",
    "NUM. FATURA", "CONVÊNIO", "NUM. REMESSA", "DATA DA REMESSA",
    "VALOR DO PROC.", "ATO", "VIA DE ACESSO", "PORTE CIRÚRGICO",
    "ACOMODAÇÃO", "PORTA DE ENTRADA",
]

ADDITIONAL_COLS_ORDER = [
    "PORTE", "CBHPM", "QTD_AUX_TABELA",
    "CIRURGIÃO", "1º AUXILIAR", "2º AUXILIAR", "3º AUXILIAR",
    "DEFLATOR", "VALOR DE REP. REGRA",
    "VALOR REP. SISHOP", "COMPLEMENTE", "DEDUÇÃO", "PROFISSIONAL",
]


def to_float_safe(x, default=0.0):
    if x is None:
        return default
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return default
    s2 = s.replace(".", "").replace(",", ".") if "," in s else s
    try:
        return float(s2)
    except:
        return default


def read_txt(uploaded_file):
    encodings = ["utf-8", "utf-8-sig", "latin-1"]
    for enc in encodings:
        try:
            df = pd.read_csv(uploaded_file, sep="\t", header=None, dtype=str,
                             encoding=enc, engine="python", keep_default_na=False)
            break
        except:
            continue

    if df.shape[1] < 25:
        for _ in range(25-df.shape[1]):
            df[df.shape[1]] = ""

    df = df.iloc[:, :25]
    df.columns = FIXED_HEADERS_A_TO_Y
    return df.astype(str)


def load_tabela(uploaded_file):
    try:
        return pd.read_excel(uploaded_file, sheet_name="TABELA", dtype=object)
    except:
        st.error("Erro ao ler TABELA.xlsx ou aba 'TABELA' inexistente.")
        return pd.DataFrame()


def build_tabela_map(tabela_df):
    if tabela_df.empty:
        return pd.DataFrame(columns=["KEY_TUSS", "PORTE"])

    key_col = tabela_df.columns[4] if len(
        tabela_df.columns) > 4 else tabela_df.columns[0]
    porte_col = tabela_df.columns[8] if len(tabela_df.columns) > 8 else None

    out = pd.DataFrame()
    out["KEY_TUSS"] = tabela_df[key_col].astype(str).str.strip()
    out["PORTE"] = tabela_df[porte_col] if porte_col else ""
    return out


def compute_preview(df):
    for col in ["CBHPM", "VIA DE ACESSO", "QUANTIDADE", "QTD_AUX_TABELA"]:
        if col not in df.columns:
            df[col] = 0

    cbhpm = df["CBHPM"].apply(lambda v: to_float_safe(v, 0))
    via = df["VIA DE ACESSO"].apply(lambda v: to_float_safe(v, 1))
    qtd = df["QUANTIDADE"].apply(lambda v: to_float_safe(v, 0))
    qtd_aux = df["QTD_AUX_TABELA"].apply(lambda v: int(to_float_safe(v, 0)))

    cir = cbhpm * via * qtd
    aux1 = cir * 0.3 * (qtd_aux >= 1)
    aux2 = cir * 0.2 * (qtd_aux >= 2)
    aux3 = aux1 * 0.1 * (qtd_aux >= 3)
    deflator = (aux1 + aux2 + aux3) * 0.2
    regra = (cir + aux1 + aux2 + aux3) - deflator

    df["CIRURGIÃO"] = cir
    df["1º AUXILIAR"] = aux1
    df["2º AUXILIAR"] = aux2
    df["3º AUXILIAR"] = aux3
    df["DEFLATOR"] = deflator
    df["VALOR DE REP. REGRA"] = regra

    return df


def build_excel(final_df, tabela_df):

    wb = Workbook()
    wb.remove(wb.active)

    ws_tab = wb.create_sheet("TABELA")
    if not tabela_df.empty:
        ws_tab.append(list(tabela_df.columns))
        for _, row in tabela_df.iterrows():
            ws_tab.append([row.get(c) for c in tabela_df.columns])
    else:
        ws_tab.append(["(Sem TABELA carregada)"])

    ws = wb.create_sheet("PROCESSADO")
    ws.append(list(final_df.columns))

    header_font = Font(bold=True)
    cols = list(final_df.columns)

    for col in range(1, ws.max_column+1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    def col_letter(name):
        return get_column_letter(cols.index(name)+1)

    numeric_zero_cols = [
        "REGISTRO", "CÓD. PRODUTO", "QUANTIDADE",
        "CÓD. TUSS", "NUM. REMESSA"
    ]

    money_two_dec_cols = [
        "CBHPM", "CIRURGIÃO", "1º AUXILIAR",
        "2º AUXILIAR", "3º AUXILIAR",
        "DEFLATOR", "VALOR DE REP. REGRA"
    ]

    for r_idx, (_, row) in enumerate(final_df.iterrows(), start=2):

        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            value = row.get(col, "")

            if col in numeric_zero_cols:
                try:
                    cell.value = float(str(value).replace(",", "."))
                except:
                    cell.value = 0
                cell.number_format = "0"

            elif col == "VALOR DO PROC.":
                cell.value = to_float_safe(value, 0)
                cell.number_format = "#,##0.00"

            elif col == "VIA DE ACESSO":
                cell.value = to_float_safe(value, 0)
                cell.number_format = "0%"

            elif col in money_two_dec_cols:
                cell.number_format = "#,##0.00"

            else:
                cell.value = value

        # ========= FÓRMULAS =========

        cod_tuss = f"{col_letter('CÓD. TUSS')}{r_idx}"
        hospital = f"{col_letter('UNIDADE')}{r_idx}"

        ws[f"{col_letter('CBHPM')}{r_idx}"] = (
            f'=IFERROR('
            f'IF({hospital}="HOSPITAL VITORIA",'
            f'VLOOKUP(VALUE({cod_tuss}),TABELA!E:Z,MATCH("VALOR H. VIT.",TABELA!$E$1:$Z$1,0),FALSE),'
            f'IF({hospital}="HOSPITAL SAMARITANO",'
            f'VLOOKUP(VALUE({cod_tuss}),TABELA!E:Z,MATCH("VALOR H. SAM.",TABELA!$E$1:$Z$1,0),FALSE),0)'
            f'),0)'
        )

        ws[f"{col_letter('QTD_AUX_TABELA')}{r_idx}"] = (
            f'=IFERROR(VLOOKUP(VALUE({cod_tuss}),TABELA!E:K,7,FALSE),0)'
        )

        # NOVOS CÁLCULOS CONFORME PRINTS

        ws[f"{col_letter('CIRURGIÃO')}{r_idx}"] = f'=IF(U{r_idx}="CIRURGIAO",AA{r_idx}*V{r_idx}*H{r_idx},0)'

        ws[f"{col_letter('1º AUXILIAR')}{r_idx}"] = f'=IF(U{r_idx}="1º AUXILIAR",IF(AB{r_idx}>=1,(AA{r_idx}*V{r_idx}*H{r_idx})*0.3,0),0)'

        ws[f"{col_letter('2º AUXILIAR')}{r_idx}"] = f'=IF(U{r_idx}="2º AUXILIAR",IF(AB{r_idx}>=2,(AA{r_idx}*V{r_idx}*H{r_idx})*0.2,0),0)'

        ws[f"{col_letter('3º AUXILIAR')}{r_idx}"] = f'=IF(U{r_idx}="3º AUXILIAR",IF(AB{r_idx}>=3,(AA{r_idx}*V{r_idx}*H{r_idx})*0.1,0),0)'

        ws[f"{col_letter('DEFLATOR')}{r_idx}"] = f"=({col_letter('1º AUXILIAR')}{r_idx}+{col_letter('2º AUXILIAR')}{r_idx}+{col_letter('3º AUXILIAR')}{r_idx})*0.2"

        ws[f"{col_letter('VALOR DE REP. REGRA')}{r_idx}"] = f"=({col_letter('CIRURGIÃO')}{r_idx}+{col_letter('1º AUXILIAR')}{r_idx}+{col_letter('2º AUXILIAR')}{r_idx}+{col_letter('3º AUXILIAR')}{r_idx})-{col_letter('DEFLATOR')}{r_idx}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


st.title("🏥 HM — Processador Hospitalar TXT → Excel")

txt_files = st.file_uploader("Envie os arquivos TXT", type=[
                             "txt"], accept_multiple_files=True)
tabela_file = st.file_uploader("Envie TABELA.xlsx", type=["xlsx"])

if st.button("🔄 Processar"):
    if not txt_files:
        st.warning("Envie ao menos um TXT.")
        st.stop()

    dfs = [read_txt(f) for f in txt_files]
    main_df = pd.concat(dfs, ignore_index=True)

    tabela_df = load_tabela(tabela_file) if tabela_file else pd.DataFrame()
    tabela_map = build_tabela_map(tabela_df)

    if not tabela_map.empty:
        main_df = main_df.merge(
            tabela_map,
            how="left",
            left_on="CÓD. TUSS",
            right_on="KEY_TUSS"
        ).drop(columns=["KEY_TUSS"])

    for col in ADDITIONAL_COLS_ORDER:
        if col not in main_df.columns:
            main_df[col] = ""

    preview_df = compute_preview(main_df.copy())

    st.session_state["final_df"] = main_df
    st.session_state["preview_df"] = preview_df
    st.session_state["tabela_df"] = tabela_df

preview_df = st.session_state.get("preview_df", pd.DataFrame())
final_df = st.session_state.get("final_df", pd.DataFrame())
tabela_df = st.session_state.get("tabela_df", pd.DataFrame())

if not preview_df.empty:
    st.subheader("Preview")
    st.dataframe(preview_df.head(200), use_container_width=True)

if not final_df.empty:
    excel_bytes = build_excel(final_df, tabela_df)
    st.download_button(
        "📥 Exportar Excel (com fórmulas)",
        data=excel_bytes,
        file_name=f"HM_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

if st.button("🧹 Limpar sessão"):
    st.session_state.clear()
    st.success("Sessão limpa.")
